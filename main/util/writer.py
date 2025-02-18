from copy import copy
from typing import Dict
import pandas as pd
from io import BytesIO

import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from main.constants import c_text
from main.data.condition_container import ConditionContainer
from main.data.data_container import DataContainer
from main.layout.layout_output_data import LayoutOutputData
from main.layout.layout_output_format_data import LayoutOutputDataFormat

class Writer:
    @classmethod
    def apply_conditional_formatting(cls, ws, fmt_condition, sheet_df, country_label, ticker_list, gap_ptr):
        """
        Apply conditional formatting based on given conditions.

        :param ws: Excel Worksheet object
        :param fmt_condition: Dictionary containing formatting conditions
        :param sheet_df: DataFrame representing the sheet data
        :param country_label: Country label key for accessing formatting conditions
        :param ticker_list: List of tickers for the given country
        :param gap_ptr: The starting row index for the country
        :return: Updated gap_ptr
        """
        if len(ticker_list) == 0:
            return gap_ptr

        # Define column lists for different metrics
        col_dict = {
            "div": [sheet_df.columns.get_loc(c_text.DIV_YIELD_TTM)],
            "capex": [
                sheet_df.columns.get_loc(c_text.CAPEX_NI_TTM),
                sheet_df.columns.get_loc(c_text.CAPEX_NI_5Y_AVG),
                sheet_df.columns.get_loc(c_text.CAPEX_NI_10Y_AVG)
            ],
            "eps": [
                sheet_df.columns.get_loc(c_text.EPS_CAGR_TTM),
                sheet_df.columns.get_loc(c_text.EPS_CAGR_3Y_TTM),
                sheet_df.columns.get_loc(c_text.EPS_CAGR_5Y_TTM),
                sheet_df.columns.get_loc(c_text.EPS_CAGR_10Y_TTM)
            ],
            "gm": [
                sheet_df.columns.get_loc(c_text.GM_LAST_Q),
                sheet_df.columns.get_loc(c_text.GM_TTM),
                sheet_df.columns.get_loc(c_text.GM_FY1),
                sheet_df.columns.get_loc(c_text.GM_FY3)
            ]
        }

        # Define colors for each metric
        fill_colors = {
            "div": "FDE9D9",
            "capex": "DAEEF3",
            "eps": "E4DFEC",
            "gm": "C5D9F1"
        }

        # Loop through each row based on tickers in the country
        for row in range(gap_ptr + 1, gap_ptr + len(ticker_list) + 1):
            for metric, cols in col_dict.items():
                for col in cols:
                    cell = ws.cell(row=row, column=col + 1)
                    if cell.value is None:
                        continue
                    if cell.value > getattr(fmt_condition[country_label], metric):
                        cell.fill = PatternFill(start_color=fill_colors[metric], end_color=fill_colors[metric], fill_type="solid")

        return gap_ptr + len(ticker_list)
    
    @classmethod
    def insert_conditional_table(cls, ws, to_insert_row, region_label, fmt_condition, border_style, c_text):
        """
        Inserts a conditional formatting table into the worksheet for a given region.
        
        Parameters:
        - ws: Worksheet object
        - to_insert_row: Row index where the table starts
        - region_label: Label for the region (e.g., "US", "CN" or "JP")
        - fmt_condition: Dictionary containing formatted condition values
        - border_style: Border style to apply
        - c_text: Object containing text constants
        """
        
        # Insert region header
        to_insert_row += 1
        cell = ws.cell(row=to_insert_row, column=1)
        cell.value = getattr(c_text, f'LABEL__{region_label}')
        
        # Insert Condition Value
        to_insert_row += 1
        value_label_cell = ws.cell(row=to_insert_row, column=1)
        value_label_cell.value = c_text.COND__VALUE
        value_label_cell.border = border_style

        div_cell = ws.cell(row=to_insert_row, column=2)
        div_cell.value = f"{c_text.COND__DIV} {fmt_condition[getattr(c_text, f'LABEL__{region_label}')].div}"
        div_cell.border = border_style

        # Insert Condition Growth
        to_insert_row += 1
        growth_label_cell = ws.cell(row=to_insert_row, column=1)
        growth_label_cell.value = c_text.COND__GROWTH
        growth_label_cell.border = border_style

        capex_cell = ws.cell(row=to_insert_row, column=2)
        capex_cell.value = f"{c_text.COND__CAPEX} {fmt_condition[getattr(c_text, f'LABEL__{region_label}')].capex}"
        capex_cell.border = border_style

        # Insert Condition EPS
        to_insert_row += 1
        eps_cell = ws.cell(row=to_insert_row, column=2)
        eps_cell.value = f"{c_text.COND__EPS} {fmt_condition[getattr(c_text, f'LABEL__{region_label}')].eps}"
        eps_cell.border = border_style

        # Insert Condition GM
        to_insert_row += 1
        gm_cell = ws.cell(row=to_insert_row, column=2)
        gm_cell.value = f"{c_text.COND__GM} {fmt_condition[getattr(c_text, f'LABEL__{region_label}')].gm}"
        gm_cell.border = border_style

        # Merge Growth label cell
        ws.merge_cells(f"A{to_insert_row - 2}:A{to_insert_row}")
        growth_label_cell.alignment = Alignment(vertical='center')

        return to_insert_row + 1

    @classmethod
    def convert_df_to_excel(self, df: pd.DataFrame, data_layout_dict: Dict[str, DataContainer], fmt_condition: Dict[str, ConditionContainer]):
        output = BytesIO()
        writer = pd.ExcelWriter(output, engine='xlsxwriter')

        unique_df: pd.DataFrame = df.copy().set_index('Ticker')
        unique_df.drop_duplicates(inplace=True)

        border_style = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for sheetname, data_container in data_layout_dict.items():
            is_value_sheet = (sheetname == c_text.LABEL__VALUE_STOCK)

            # If sheet no content, skip
            if data_container.is_empty():
                continue

            sheet_df_ls = []
            if len(data_container.us_ticker_ls) > 0:
                sheet_df_ls.append(unique_df.loc[data_container.us_ticker_ls].reset_index())
            
            if len(data_container.cn_ticker_ls) > 0:
                sheet_df_ls.append(unique_df.loc[data_container.cn_ticker_ls].reset_index())
            
            if len(data_container.jp_ticker_ls) > 0:
                sheet_df_ls.append(unique_df.loc[data_container.jp_ticker_ls].reset_index())
            
            sheet_df = pd.concat(sheet_df_ls)
            sheet_df.reset_index(drop=True, inplace=True)

            # If is value stock, reorder the columns
            if is_value_sheet:
                sheet_df = sheet_df[LayoutOutputData.col_value_order]
            else:
                sheet_df = sheet_df[LayoutOutputData.col_order]

            sheet_df.to_excel(writer, sheet_name=sheetname, index=False)
            
            # Access the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets[sheetname]

            percent_format = workbook.add_format({'num_format': '0.0%'})
            one_decimal_format = workbook.add_format({'num_format': '#,##0.0'})
            three_decimal_format = workbook.add_format({'num_format': '#,##0.000'})
            
            if LayoutOutputDataFormat.pct_col_ls:
                for col in LayoutOutputDataFormat.pct_col_ls:
                    if col in sheet_df.columns:
                        col_idx = sheet_df.columns.get_loc(col)
                        worksheet.set_column(col_idx, col_idx, None, percent_format)

            if LayoutOutputDataFormat.num_one_decim_col_ls:
                for col in LayoutOutputDataFormat.num_one_decim_col_ls:
                    if col in sheet_df.columns:
                        col_idx = sheet_df.columns.get_loc(col)
                        worksheet.set_column(col_idx, col_idx, None, one_decimal_format)

            if LayoutOutputDataFormat.num_three_decim_col_ls:
                for col in LayoutOutputDataFormat.num_three_decim_col_ls:
                    if col in sheet_df.columns:
                        col_idx = sheet_df.columns.get_loc(col)
                        worksheet.set_column(col_idx, col_idx, None, three_decimal_format)

        writer.close()

        # More customise formatting
        output.seek(0)
        wb = load_workbook(output)
        for sheetname, data_container in data_layout_dict.items():
            is_value_sheet = (sheetname == c_text.LABEL__VALUE_STOCK)
            col_order_ls = LayoutOutputData.col_value_order if is_value_sheet else LayoutOutputData.col_order

            if data_container.is_empty():
                continue

            ws = wb[sheetname]

            # Set the sheet_df back with the values in excel
            sheet_df = pd.DataFrame(ws.values)
            sheet_df.columns = sheet_df.iloc[0]  # Set first row as column names
            sheet_df = sheet_df[1:]

            # Pre-fill white content
            for row in ws.iter_rows(min_row=1, max_row=2_000, min_col=1, max_col=100):
                for cell in row:
                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            # Total column length
            tot_col_len = len(col_order_ls)

            # Apply the border to all used cells
            for row in ws.iter_rows(min_row=1, max_row=len(data_container.master_ticker_ls) + 1, min_col=1, max_col=tot_col_len):
                for cell in row:
                    cell.border = border_style

            # Header
            header_row = 1
            header_col = 1

            for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=tot_col_len):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

            # Set column default width
            for c in range(1, tot_col_len + 1):
                col_letter = get_column_letter(c)
                ws.column_dimensions[col_letter].width = 9.0

            # Specific width
            ws.column_dimensions[get_column_letter(1)].width = 20
            ws.column_dimensions[get_column_letter(2)].width = 20
            ws.column_dimensions[get_column_letter(3)].width = 20
            dt_col_ls = [c_text.NEXT_EARN_DATE, c_text.BEAT_EST_LAST_UPDATE, c_text.LAST_EX_DIV_DT]
            for dt_col in dt_col_ls:
                ws.column_dimensions[get_column_letter(sheet_df.columns.get_loc(dt_col) + 1)].width = 11.5
            
            ws.row_dimensions[1].height = 80.0

            # Alignment
            center_align_ls = [sheet_df.columns.get_loc(c) for c in dt_col_ls + [c_text.CCY]]
            for col in center_align_ls:
                col += 1
                for row in range(2, len(data_container.master_ticker_ls) + 2):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = Alignment(wrap_text=True, horizontal='center')
            
            right_align_ls = LayoutOutputDataFormat.txt_col_ls.copy()
            right_align_ls = list(filter(lambda x: x in col_order_ls, right_align_ls))
            right_align_ls = [sheet_df.columns.get_loc(c) for c in right_align_ls]
            for col in right_align_ls:
                col += 1
                for row in range(2, len(data_container.master_ticker_ls) + 2):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = Alignment(wrap_text=True, horizontal='right')
            
            # Copy and paste the header and rename it
            out_header_ls = []
            term = sheetname.split(' ')[0]
            
            if len(data_container.us_ticker_ls) > 0:
                out_header_ls.append([f'{c_text.LABEL__US} {term}', len(data_container.us_ticker_ls)])
            if len(data_container.cn_ticker_ls) > 0:
                out_header_ls.append([f'{c_text.LABEL__CN} {term}', len(data_container.cn_ticker_ls)])
            if len(data_container.jp_ticker_ls) > 0:
                out_header_ls.append([f'{c_text.LABEL__JP} {term}', len(data_container.jp_ticker_ls)])

            # Conditional formatting
            gap_ptr = 1
            gap_ptr = self.apply_conditional_formatting(ws, fmt_condition, sheet_df, c_text.LABEL__US, data_container.us_ticker_ls, gap_ptr)
            gap_ptr = self.apply_conditional_formatting(ws, fmt_condition, sheet_df, c_text.LABEL__CN, data_container.cn_ticker_ls, gap_ptr)
            gap_ptr = self.apply_conditional_formatting(ws, fmt_condition, sheet_df, c_text.LABEL__JP, data_container.jp_ticker_ls, gap_ptr)

            # Insert the header to rest of the country
            to_insert_row = 0
            for i, (label, tot_len) in enumerate(out_header_ls):
                if i == 0:
                    header_cell = ws.cell(row=header_row, column=header_col)
                    header_cell.value = label
                    to_insert_row += tot_len + 2
                
                    continue
                
                # Insert breaker
                ws.insert_rows(to_insert_row, 2)

                # Fill up empty cells
                for row in ws.iter_rows(min_row=to_insert_row, max_row=to_insert_row, min_col=1, max_col=100):
                    for cell in row:
                        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                to_insert_row += 1  # Move the header pointer

                # Add the new header to each group and copy the header from row 1
                for col in range(1, ws.max_column + 1):
                    cell_source = ws.cell(row=header_row, column=col)
                    cell_target = ws.cell(row=to_insert_row, column=col)

                    cell_target.value = cell_source.value

                    if cell_source.has_style:
                        cell_target.font = Font(
                            name=cell_source.font.name,
                            size=cell_source.font.size,
                            bold=cell_source.font.bold,
                            italic=cell_source.font.italic,
                            underline=cell_source.font.underline,
                            color=cell_source.font.color
                        )

                        cell_target.fill = PatternFill(
                            fill_type=cell_source.fill.fill_type,
                            start_color=cell_source.fill.start_color,
                            end_color=cell_source.fill.end_color
                        )

                        cell_target.border = Border(
                            left=cell_source.border.left,
                            right=cell_source.border.right,
                            top=cell_source.border.top,
                            bottom=cell_source.border.bottom
                        )

                        cell_target.alignment = Alignment(
                            horizontal=cell_source.alignment.horizontal,
                            vertical=cell_source.alignment.vertical,
                            wrap_text=cell_source.alignment.wrap_text
                        )

                        cell_target.number_format = cell_source.number_format
                    
                    ws.row_dimensions[to_insert_row].height = ws.row_dimensions[header_row].height

                    header_cell = ws.cell(row=to_insert_row, column=header_col)
                    header_cell.value = label
                
                to_insert_row += tot_len + 1

            # Insert the conditional formatting table
            to_insert_row += 2
            cond_cell_header = ws.cell(row=to_insert_row, column=1)
            cond_cell_header.value = c_text.COND
            cond_cell_header.fill = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type="solid")
            cond_cell_header.font = Font(bold=True)
            
            next_cond_cell_header = ws.cell(row=to_insert_row, column=2)
            next_cond_cell_header.fill = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type="solid")

            if len(data_container.us_ticker_ls) > 0:
                to_insert_row = self.insert_conditional_table(ws, to_insert_row, "US", fmt_condition, border_style, c_text)

            if len(data_container.cn_ticker_ls) > 0:
                to_insert_row = self.insert_conditional_table(ws, to_insert_row, "CN", fmt_condition, border_style, c_text)

            if len(data_container.jp_ticker_ls) > 0:
                to_insert_row = self.insert_conditional_table(ws, to_insert_row, "JP", fmt_condition, border_style, c_text)
        
                
        # Save the modified Excel file back to BytesIO
        final_output = BytesIO()
        wb.save(final_output)

        # Get processed data
        processed_data = final_output.getvalue()
 
        return processed_data
    